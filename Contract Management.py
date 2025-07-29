import streamlit as st
import pandas as pd
import plotly.express as px
import os
import io
from datetime import datetime
import streamlit_authenticator as stauth
import json
import requests

FILE_NAME = "contracts.xlsx"
SHEETBEST_API_URL = "https://api.sheetbest.com/sheets/befd2b02-ac57-42fa-93be-1d37ccd5291d"

st.set_page_config(page_title="Qu·∫£n l√Ω h·ª£p ƒë·ªìng", layout="wide")

# === Google Sheets ===
def save_to_google_sheets(df):
    try:
        requests.delete(SHEETBEST_API_URL)
        for record in df.to_dict(orient="records"):
            requests.post(SHEETBEST_API_URL, json=record)
    except Exception as e:
        st.warning(f"Kh√¥ng th·ªÉ ƒë·ªìng b·ªô l√™n Sheet.best: {e}")

def load_from_google_sheets():
    try:
        res = requests.get(SHEETBEST_API_URL)
        res.raise_for_status()
        return pd.DataFrame(res.json())
    except Exception as e:
        st.warning(f"Kh√¥ng th·ªÉ t·∫£i d·ªØ li·ªáu t·ª´ Sheet.best: {e}")
        return pd.DataFrame()

# === User Authentication ===
def load_users():
    # S·ª≠ d·ª•ng m·∫≠t kh·∫©u ƒë√£ hash s·∫µn cho '123456'
    hashed_passwords = ["$2b$12$KIXt87YOD41xZtMdpo97fOVJrNOxZbDTRZKFa6xB6KOe4a6DFi2lW"]
    return {
        "usernames": {
            "admin": {
                "name": "Admin",
                "password": hashed_passwords[0]
            }
        }
    }

credentials = load_users()
authenticator = stauth.Authenticate(
    credentials,
    "contract_app",
    "auth_token",
    cookie_expiry_days=1
)
auth_status = True
name = "Admin"
username = "admin"

# X√ìA X√ÅC TH·ª∞C ƒë·ªÉ xem giao di·ªán
st.sidebar.success(f"‚úÖ Xin ch√†o, {name}")
st.title("üìã Qu·∫£n l√Ω H·ª£p ƒë·ªìng & ƒê∆°n h√†ng")

with st.sidebar.expander("üè¢ Th√¥ng tin doanh nghi·ªáp"):
    company_name = st.text_input("T√™n doanh nghi·ªáp", "C√¥ng ty TNHH ABC")
    logo_file = st.file_uploader("T·∫£i l√™n logo", type=["png", "jpg", "jpeg"])
    if logo_file:
        st.image(logo_file, use_column_width=True)
    st.markdown(f"**T√™n doanh nghi·ªáp:** {company_name}")

    df = load_from_google_sheets()

    # L√†m s·∫°ch d·ªØ li·ªáu ng√†y h√≥a ƒë∆°n n·∫øu c√≥
    if "Ng√†y h√≥a ƒë∆°n" in df.columns:
        df["Ng√†y h√≥a ƒë∆°n"] = pd.to_datetime(df["Ng√†y h√≥a ƒë∆°n"], errors="coerce")

    # T√≠nh to√°n c√°c ƒë·ª£t thanh to√°n v√† gi√° tr·ªã c√≤n l·∫°i n·∫øu c√≥ c·ªôt L·ªãch s·ª≠ thanh to√°n
    if "L·ªãch s·ª≠ thanh to√°n" in df.columns:
        def parse_ltt(x):
            if pd.isna(x) or not isinstance(x, str):
                return 0
            parts = x.split(";")
            return sum([float(p.split("|")[1]) for p in parts if "|" in p])

        df["T·ªïng ƒë√£ thanh to√°n"] = df["L·ªãch s·ª≠ thanh to√°n"].apply(parse_ltt)
        df["C√≤n l·∫°i"] = df["Gi√° tr·ªã quy·∫øt to√°n"].fillna(0) - df["T·ªïng ƒë√£ thanh to√°n"].fillna(0)

    if st.button("‚ûï Th√™m h·ª£p ƒë·ªìng m·ªõi"):
        with st.form("form_them_hop_dong", clear_on_submit=True):
            st.subheader("üìù Nh·∫≠p th√¥ng tin h·ª£p ƒë·ªìng m·ªõi")
            ma_hd = st.text_input("M√£ h·ª£p ƒë·ªìng")
            khach_hang = st.text_input("Kh√°ch h√†ng")
            ngay_ky = st.date_input("Ng√†y k√Ω")
            gt_quyet_toan = st.number_input("Gi√° tr·ªã quy·∫øt to√°n", min_value=0.0, step=100000.0)
            trang_thai_hd = st.selectbox("Tr·∫°ng th√°i h·ª£p ƒë·ªìng", ["ƒêang th·ª±c hi·ªán", "Ho√†n th√†nh", "H·ªßy"])
            trang_thai_hdong = st.selectbox("Tr·∫°ng th√°i h√≥a ƒë∆°n", ["Ch∆∞a xu·∫•t", "ƒê√£ xu·∫•t"])
            so_hoa_don = st.text_input("S·ªë h√≥a ƒë∆°n")
            ngay_hoa_don = st.date_input("Ng√†y h√≥a ƒë∆°n") if trang_thai_hdong == "ƒê√£ xu·∫•t" else ""
            lich_su_tt = st.text_area("L·ªãch s·ª≠ thanh to√°n", help="Nh·∫≠p d·∫°ng: Ng√†y|Gi√° tr·ªã;Ng√†y|Gi√° tr·ªã")
            submitted = st.form_submit_button("L∆∞u h·ª£p ƒë·ªìng")
            if submitted:
                new_row = {
                    "M√£ h·ª£p ƒë·ªìng": ma_hd,
                    "Kh√°ch h√†ng": khach_hang,
                    "Ng√†y k√Ω": ngay_ky.strftime("%Y-%m-%d"),
                    "Gi√° tr·ªã quy·∫øt to√°n": gt_quyet_toan,
                    "Tr·∫°ng th√°i h·ª£p ƒë·ªìng": trang_thai_hd,
                    "Tr·∫°ng th√°i h√≥a ƒë∆°n": trang_thai_hdong,
                    "S·ªë h√≥a ƒë∆°n": so_hoa_don,
                    "Ng√†y h√≥a ƒë∆°n": ngay_hoa_don.strftime("%Y-%m-%d") if ngay_hoa_don else "",
                    "L·ªãch s·ª≠ thanh to√°n": lich_su_tt
                }
                df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
                save_to_google_sheets(df)
                st.success("‚úÖ H·ª£p ƒë·ªìng m·ªõi ƒë√£ ƒë∆∞·ª£c l∆∞u!")

    if df.empty:
        st.info("Ch∆∞a c√≥ d·ªØ li·ªáu h·ª£p ƒë·ªìng.")
    else:
        # L√†m s·∫°ch d·ªØ li·ªáu ng√†y
        df["Ng√†y k√Ω"] = pd.to_datetime(df["Ng√†y k√Ω"], errors="coerce")
        df["NƒÉm"] = df["Ng√†y k√Ω"].dt.year
        df["Th√°ng"] = df["Ng√†y k√Ω"].dt.month
        df["Qu√Ω"] = df["Ng√†y k√Ω"].dt.quarter

        st.subheader("üìë Danh s√°ch h·ª£p ƒë·ªìng")
        # Hi·ªÉn th·ªã th√™m c·ªôt tr·∫°ng th√°i h·ª£p ƒë·ªìng, tr·∫°ng th√°i h√≥a ƒë∆°n, t·ªïng ƒë√£ thanh to√°n, c√≤n l·∫°i
        display_cols = [
            "M√£ h·ª£p ƒë·ªìng", "Kh√°ch h√†ng", "Ng√†y k√Ω", "Gi√° tr·ªã quy·∫øt to√°n",
            "Tr·∫°ng th√°i h·ª£p ƒë·ªìng", "Tr·∫°ng th√°i h√≥a ƒë∆°n", "Ng√†y h√≥a ƒë∆°n", "S·ªë h√≥a ƒë∆°n",
            "T·ªïng ƒë√£ thanh to√°n", "C√≤n l·∫°i"
        ]
        for col in display_cols:
            if col not in df.columns:
                df[col] = None
        st.dataframe(df[display_cols], use_container_width=True)

        # Ch·ªânh s·ª≠a h·ª£p ƒë·ªìng theo m√£
        st.subheader("‚úèÔ∏è C·∫≠p nh·∫≠t thanh to√°n v√† gi√° tr·ªã quy·∫øt to√°n")
        selected_hd = st.selectbox("Ch·ªçn h·ª£p ƒë·ªìng ƒë·ªÉ ch·ªânh s·ª≠a", df["M√£ h·ª£p ƒë·ªìng"].dropna().unique())
        if selected_hd:
            row = df[df["M√£ h·ª£p ƒë·ªìng"] == selected_hd].iloc[0]
            with st.form("form_sua_hd"):
                gt_moi = st.number_input("C·∫≠p nh·∫≠t gi√° tr·ªã quy·∫øt to√°n", value=float(row["Gi√° tr·ªã quy·∫øt to√°n"]))
                lich_su_moi = st.text_area("C·∫≠p nh·∫≠t l·ªãch s·ª≠ thanh to√°n", value=row.get("L·ªãch s·ª≠ thanh to√°n", ""),
                                          help="Nh·∫≠p d·∫°ng: Ng√†y|Gi√° tr·ªã;Ng√†y|Gi√° tr·ªã")
                submit_sua = st.form_submit_button("üíæ C·∫≠p nh·∫≠t")
                if submit_sua:
                    df.loc[df["M√£ h·ª£p ƒë·ªìng"] == selected_hd, "Gi√° tr·ªã quy·∫øt to√°n"] = gt_moi
                    df.loc[df["M√£ h·ª£p ƒë·ªìng"] == selected_hd, "L·ªãch s·ª≠ thanh to√°n"] = lich_su_moi
                    save_to_google_sheets(df)
                    st.success("‚úÖ ƒê√£ c·∫≠p nh·∫≠t h·ª£p ƒë·ªìng!")
        col_filter1, col_filter2 = st.columns(2)
        with col_filter1:
            selected_kh = st.selectbox("üë§ L·ªçc theo kh√°ch h√†ng", ["T·∫•t c·∫£"] + sorted(df["Kh√°ch h√†ng"].dropna().unique()))
        with col_filter2:
            selected_invoice = "T·∫•t c·∫£"
            if "Tr·∫°ng th√°i h√≥a ƒë∆°n" in df.columns:
                invoice_options = ["T·∫•t c·∫£"] + sorted(df["Tr·∫°ng th√°i h√≥a ƒë∆°n"].dropna().unique())
                selected_invoice = st.selectbox("üßæ L·ªçc theo tr·∫°ng th√°i h√≥a ƒë∆°n", invoice_options)
                if selected_invoice != "T·∫•t c·∫£":
                    df = df[df["Tr·∫°ng th√°i h√≥a ƒë∆°n"] == selected_invoice]

        if selected_kh != "T·∫•t c·∫£":
            df = df[df["Kh√°ch h√†ng"] == selected_kh]
        if selected_invoice != "T·∫•t c·∫£":
            df = df[df["Tr·∫°ng th√°i h√≥a ƒë∆°n"] == selected_invoice]
        year_options = ["T·∫•t c·∫£"] + sorted(df["NƒÉm"].dropna().unique().astype(int).tolist())
        selected_year = st.selectbox("üìÜ L·ªçc theo nƒÉm", year_options)

        if selected_year != "T·∫•t c·∫£":
            df = df[df["NƒÉm"] == int(selected_year)]

        st.dataframe(df, use_container_width=True)

        st.subheader("üìä Th·ªëng k√™ doanh thu")
        col1, col2 = st.columns(2)

        with col1:
            month_stat = df.groupby("Th√°ng")["Gi√° tr·ªã quy·∫øt to√°n"].sum().reset_index()
            fig1 = px.bar(month_stat, x="Th√°ng", y="Gi√° tr·ªã quy·∫øt to√°n", title="Doanh thu theo th√°ng")
            st.plotly_chart(fig1, use_container_width=True)

        with col2:
            quarter_stat = df.groupby("Qu√Ω")["Gi√° tr·ªã quy·∫øt to√°n"].sum().reset_index()
            fig2 = px.pie(quarter_stat, names="Qu√Ω", values="Gi√° tr·ªã quy·∫øt to√°n", title="T·ª∑ l·ªá doanh thu theo qu√Ω")
            st.plotly_chart(fig2, use_container_width=True)

        st.subheader("üë• Doanh thu theo kh√°ch h√†ng")

        st.subheader("üì§ Xu·∫•t b√°o c√°o theo t·ª´ng kh√°ch h√†ng")
        selected_kh_xuat = st.selectbox("Ch·ªçn kh√°ch h√†ng ƒë·ªÉ xu·∫•t b√°o c√°o", sorted(df["Kh√°ch h√†ng"].dropna().unique()))
        df_kh = df[df["Kh√°ch h√†ng"] == selected_kh_xuat]

        buffer_kh = io.BytesIO()
        with pd.ExcelWriter(buffer_kh, engine="openpyxl") as writer:
            df_kh.to_excel(writer, index=False, sheet_name="H·ª£p ƒë·ªìng")
            ws = writer.sheets["H·ª£p ƒë·ªìng"]
            if company_name:
                ws.insert_rows(0)
                ws.cell(row=1, column=1).value = f"Doanh nghi·ªáp: {company_name}"
            if logo_file:
                from openpyxl.drawing.image import Image as XLImage
                from PIL import Image
                import tempfile
                with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp_img:
                    img = Image.open(logo_file)
                    img.save(tmp_img.name)
                    xl_img = XLImage(tmp_img.name)
                    xl_img.width = 150
                    xl_img.height = 60
                    ws.add_image(xl_img, "F1")
        st.download_button(
    label=f"üìÑ T·∫£i b√°o c√°o c·ªßa {selected_kh_xuat}",
    data=buffer_kh.getvalue(),
    file_name=f"bao_cao_{selected_kh_xuat}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
        st.subheader("üì¶ Xu·∫•t t·∫•t c·∫£ b√°o c√°o theo t·ª´ng kh√°ch h√†ng")
        if st.button("üìÅ T·∫£i t·∫•t c·∫£ b√°o c√°o"):
            from zipfile import ZipFile
            zip_buffer = io.BytesIO()
            with ZipFile(zip_buffer, "w") as zip_file:
                for kh in df["Kh√°ch h√†ng"].dropna().unique():
                    df_kh_all = df[df["Kh√°ch h√†ng"] == kh]
                    temp_buffer = io.BytesIO()
                    with pd.ExcelWriter(temp_buffer, engine="openpyxl") as writer:
                        df_kh_all.to_excel(writer, index=False, sheet_name="H·ª£p ƒë·ªìng")
                        ws = writer.sheets["H·ª£p ƒë·ªìng"]
                        if company_name:
                            ws.insert_rows(0)
                            ws.cell(row=1, column=1).value = f"Doanh nghi·ªáp: {company_name}"
                        if logo_file:
                            from openpyxl.drawing.image import Image as XLImage
                            from PIL import Image
                            import tempfile
                            with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp_img:
                                img = Image.open(logo_file)
                                img.save(tmp_img.name)
                                xl_img = XLImage(tmp_img.name)
                                xl_img.width = 150
                                xl_img.height = 60
                                ws.add_image(xl_img, "F1")
                    zip_file.writestr(f"{kh}_hop_dong.xlsx", temp_buffer.getvalue())
            st.download_button(
                label="üì¶ T·∫£i t·∫•t c·∫£ b√°o c√°o kh√°ch h√†ng (ZIP)",
                data=zip_buffer.getvalue(),
                file_name="bao_cao_tat_ca_khach_hang.zip",
                mime="application/zip"
            )
        kh_stat = df.groupby("Kh√°ch h√†ng")["Gi√° tr·ªã quy·∫øt to√°n"].sum().reset_index().sort_values(by="Gi√° tr·ªã quy·∫øt to√°n", ascending=False)
        fig3 = px.bar(kh_stat, x="Gi√° tr·ªã quy·∫øt to√°n", y="Kh√°ch h√†ng", orientation="h")
        st.plotly_chart(fig3, use_container_width=True)

        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            kh_stat.to_excel(writer, index=False, sheet_name="Th·ªëng k√™")
            ws = writer.sheets["Th·ªëng k√™"]

            # Ghi th√¥ng tin doanh nghi·ªáp ·ªü d√≤ng ƒë·∫ßu ti√™n n·∫øu c√≥ t√™n
            if company_name:
                ws.insert_rows(0)
                ws.cell(row=1, column=1).value = f"Doanh nghi·ªáp: {company_name}"

            # Th√™m logo n·∫øu c√≥
            if logo_file:
                from openpyxl.drawing.image import Image as XLImage
                from PIL import Image
                import tempfile

                with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp_img:
                    img = Image.open(logo_file)
                    img.save(tmp_img.name)
                    xl_img = XLImage(tmp_img.name)
                    xl_img.width = 150
                    xl_img.height = 60
                    ws.add_image(xl_img, "F1")
                ws.insert_rows(0)
                ws.cell(row=1, column=1).value = f"Doanh nghi·ªáp: {company_name}"

        st.download_button(
            label="üì• T·∫£i th·ªëng k√™ theo kh√°ch h√†ng",
            data=buffer.getvalue(),
            file_name="doanh_thu_theo_khach_hang.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

# elif auth_status is False:
  #  st.error("‚ùå Sai t√™n ƒëƒÉng nh·∫≠p ho·∫∑c m·∫≠t kh·∫©u")
# elif auth_status is None:
  #  st.warning("üîí Vui l√≤ng ƒëƒÉng nh·∫≠p ƒë·ªÉ ti·∫øp t·ª•c")
